from typing import Any, Text, Dict, List
import json

from sqlalchemy import true
from rasa_sdk import Action, Tracker
from rasa_sdk.events import (
    SlotSet,
    UserUtteranceReverted,
    ConversationPaused,
    EventType,
)
from rasa_sdk.types import DomainDict
from rasa_sdk.executor import CollectingDispatcher
import xlrd
import pandas as pd
import os
from openpyxl import Workbook, load_workbook

USER_INTENT_OUT_OF_SCOPE = "out_of_scope"
INTENT_DESCRIPTION_MAPPING_PATH = "actions/intent_description_mapping.csv"
ACTION_DEFAULT_ASK_AFFIRMATION_NAME = "action_default_ask_affirmation"
class ActionGetFeedback(Action):

    def name(self) -> Text:
        return "action_get_feedback"
    
    async def run(
        self, dispatcher, tracker: Tracker, domain: Dict[Text, Any],
    ) -> List[Dict[Text, Any]]:
        # user_name = tracker.get_slot('name')
        # print(user_name)
        cat = tracker.get_slot("feedback_name")
        if(cat == "General"):
            dispatcher.utter_message(text="Please fill the form at this [link](https://docs.google.com/forms/d/e/1FAIpQLSf2mX_PTth-HYCHuWuHLWBSiWJZxh5SpLRgPnBI7_OZr5o5uQ/viewform?usp=sf_link). Thanks for your feedback😊 !!")
        elif(cat == "College"):
            dispatcher.utter_message(text="Please fill the form at this [link](https://docs.google.com/forms/d/e/1FAIpQLSf2mX_PTth-HYCHuWuHLWBSiWJZxh5SpLRgPnBI7_OZr5o5uQ/viewform?usp=sf_link). Thanks for your feedback😊 !!")
        elif(cat == "Library"):
            dispatcher.utter_message(text="Please fill the form at this [link](https://docs.google.com/forms/d/e/1FAIpQLSf2mX_PTth-HYCHuWuHLWBSiWJZxh5SpLRgPnBI7_OZr5o5uQ/viewform?usp=sf_link). Thanks for your feedback😊 !!")
        elif(cat == "Clubs"):
            dispatcher.utter_message(text="Please fill the form at this [link](https://docs.google.com/forms/d/e/1FAIpQLSf2mX_PTth-HYCHuWuHLWBSiWJZxh5SpLRgPnBI7_OZr5o5uQ/viewform?usp=sf_link). Thanks for your feedback😊 !!")
        elif(cat == "Events"):
            dispatcher.utter_message(text="Please fill the form at this [link](https://docs.google.com/forms/d/e/1FAIpQLSf2mX_PTth-HYCHuWuHLWBSiWJZxh5SpLRgPnBI7_OZr5o5uQ/viewform?usp=sf_link). Thanks for your feedback😊 !!")
        elif(cat == "Others"):
            dispatcher.utter_message(text="Please fill the form at this [link](https://docs.google.com/forms/d/e/1FAIpQLSf2mX_PTth-HYCHuWuHLWBSiWJZxh5SpLRgPnBI7_OZr5o5uQ/viewform?usp=sf_link). Thanks for your feedback😊 !!")

        # feedback = tracker.latest_message['text']
        # data = {"Name":[user_name],"Feedback":[feedback],"Category":[sheet_name]}
        # print(data)
        # df = pd.DataFrame(data)
        # print(df)
        # excel_name = os.path.join(os.getcwd(), os.path.relpath("actions/Resources/Feedback.xlsx"))
        

        # wb= Workbook()
        # ws=wb.active
        # with pd.ExcelWriter(excel_name, engine="openpyxl") as writer:
        #     writer.book=wb
        #     writer.sheets = dict((ws.title, ws) for ws in wb.worksheets)
        #     #useful code
        #     df.to_excel(writer, sheet_name= sheet_name, index = False)
        #     writer.save()
        
        # book = load_workbook(excel_name)
        # writer = pd.ExcelWriter(excel_name, engine='openpyxl')
        # writer.book = book
        # writer.sheets = {ws.title: ws for ws in book.worksheets}

        # for sheetname in writer.sheets:
        #     df.to_excel(writer,sheet_name=sheetname, startrow=writer.sheets[sheetname].max_row, index = False,header= False)

        # writer.save()

        # with pd.ExcelWriter(excel_name,engine='xlsxwriter') as writer:
        #     columns = []
        #     for k, v in data.items():
        #         columns.append(k)
        #     df = pd.DataFrame(data, index=[0])
        #     df_source = None
        #     if os.path.exists(excel_name):
        #         df_source = pd.DataFrame(pd.read_excel(excel_name, sheet_name=sheet_name, engine='openpyxl'))
        #     if df_source is not None:
        #         df_dest = df_source.append(df)
        #     else:
        #         df_dest = df
        #     df_dest.to_excel(writer, sheet_name=sheet_name, index = False, columns=columns)
        
          
        #dispatcher.utter_message(text="Thanks for your feedback😊 !!")
        
        return []
    

#locally queries the database for clubs info
class ActionTellClubInfo(Action):

    def name(self) -> Text:
        return "action_tell_club_info"

    async def run(
        self, dispatcher, tracker: Tracker, domain: Dict[Text, Any],
    ) -> List[Dict[Text, Any]]:
        club_name = next(tracker.get_latest_entity_values("club_name"), None)
        
        loc = os.path.join(os.getcwd(), os.path.relpath("actions/Resources/club_details.json"))

        data = pd.read_json(loc)  
        data = pd.DataFrame(data)
        result_data = data.query("Name == @club_name")
        if not result_data.empty:
            msg = "Club Name : "+club_name+"\nClub Lead : "+result_data["Club_Lead"].iloc[0]+ "\nClub President : "+result_data["Club_President"].iloc[0]+"\nClub Description : "+result_data["Club_Description"].iloc[0]+"\nInstagram Handle : https://instagram.com/"+result_data["Insta_handle"].iloc[0]
        else:
            msg = f"The club you are looking for doesn't seem to exist. Could you please check again"
          
        dispatcher.utter_message(text=msg)
        
        return []

class ActionTellEventInfo(Action):

    def name(self) -> Text:
        return "action_tell_event_info"

    async def run(
        self, dispatcher, tracker: Tracker, domain: Dict[Text, Any],
    ) -> List[Dict[Text, Any]]:
        event_name = next(tracker.get_latest_entity_values("event_name"), None)
        
        loc = os.path.join(os.getcwd(), os.path.relpath("actions/Resources/event_details.json"))

        data = pd.read_json(loc)  
        data = pd.DataFrame(data)
        result_data = data.query("Event_Name == @event_name")
        if not result_data.empty:
            msg = "Event Name : "+event_name+"\nWho to contact : "+result_data["Who_to_contact"].iloc[0]+ "\nEvent Details : "+result_data["What_is_the_event"].iloc[0]
        else:
            msg = f"The event you are looking for doesn't seem to exist. Could you please check again"
          
        dispatcher.utter_message(text=msg)
        
        return []
# class ActionButtonForClubInfo(Action):

#     def name(self) -> Text:
#         return "action_button_for_club_info"

#     def run(self, dispatcher, 
#             tracker: Tracker, 
#             domain: Dict[Text, Any],) -> List[Dict[Text, Any]]:
        
#         buttons = [
#             {"payload":"/yes_club_info[club_info_choice](yes)", "title":"I want to know about existing clubs"},
#             {"payload":"/no_club_info[club_info_choice](no)", "title":"Not needed"},
#         ]
#         dispatcher.utter_message("Would you like to know more about the clubs already present at BPDC?", buttons = buttons)
        
#         return []

class ActionTellStudGpa(Action):

    def name(self) -> Text:
        return "action_tell_stud_gpa"

    async def run(
        self, dispatcher, tracker: Tracker, domain: Dict[Text, Any],
    ) -> List[Dict[Text, Any]]:
        stud_id = tracker.get_slot("stud_id")
        stud_pwd = tracker.get_slot("stud_pwd")
        
        loc = os.path.join(os.getcwd(), os.path.basename("./DB.xlsx"))
        data = pd.read_excel(loc, sheet_name="Auth")  
        data = pd.DataFrame(data)
        result_data = data.query("ID == @stud_id")
        check_pwd = result_data.iloc[0]["Password"]
        msg = ""
        if not result_data.empty:
            if str(check_pwd) == str(stud_pwd):
                data1 = pd.read_excel(loc, sheet_name="GPA")  
                data1 = pd.DataFrame(data1)
                result_data1 = data1.query("ID == @stud_id")
                msg = "Your GPA is : "+str(result_data1.iloc[0]["GPA"])
            else:
                msg = "Your password does not match. Sorry!!"
            
        else:
            msg = f"The ID does not exist in our database!! Please check it again."

        
          
        dispatcher.utter_message(text=msg)
        
        return []

# class ActionGreetUser(Action):
#     """Greets the user with/without privacy policy"""

#     def name(self) -> Text:
#         return "action_greet_user"

#     def run(
#         self,
#         dispatcher: CollectingDispatcher,
#         tracker: Tracker,
#         domain: DomainDict,
#     ) -> List[EventType]:
#         intent = tracker.latest_message["intent"].get("name")
#         # shown_privacy = tracker.get_slot("shown_privacy")
#         shown_privacy = true
#         name_entity = next(tracker.get_latest_entity_values("name"), None)
#         if intent == "greet" or (intent == "enter_data" and name_entity):
#             if shown_privacy and name_entity and name_entity.lower() != "sara":
#                 dispatcher.utter_message(response="utter_greet_name", name=name_entity)
#                 return []
#             elif shown_privacy:
#                 dispatcher.utter_message(response="utter_greet_noname")
#                 return []
#             else:
#                 dispatcher.utter_message(response="utter_greet")
#                 dispatcher.utter_message(response="utter_inform_privacypolicy")
#                 return [SlotSet("shown_privacy", True)]
#         return []

class ActionRestartWithButton(Action):
    def name(self) -> Text:
        return "action_restart_with_button"

    def run(
        self,
        dispatcher: CollectingDispatcher,
        tracker: Tracker,
        domain: DomainDict,
    ) -> None:

        dispatcher.utter_message(response="utter_restart_with_button")

class ActionTellClubChoices(Action):
    def name(self) -> Text:
        return "action_tell_club_choices"

    def run(
        self,
        dispatcher: CollectingDispatcher,
        tracker: Tracker,
        domain: DomainDict,
    ) -> None:
        buttons = []
        loc = os.path.join(os.getcwd(), os.path.relpath("actions/Resources/club_details.json"))
        data = pd.read_json(loc)  
        data = pd.DataFrame(data)
        entity_name = "club_name"
        for i in range(len(data["Name"])):
            #############################################
            e_name = "club_name"
            e_value = data["Name"][i]
            buttons.append(
                        {"title": e_value, "payload": "/club_choice "+json.dumps({e_name:e_value})}
                    )
            
        dispatcher.utter_message(response="utter_club_name_details", buttons=buttons, button_type = "vertical")

class ActionTellEventChoices(Action):
    def name(self) -> Text:
        return "action_tell_event_choices"

    def run(
        self,
        dispatcher: CollectingDispatcher,
        tracker: Tracker,
        domain: DomainDict,
    ) -> None:
        buttons = []
        loc = os.path.join(os.getcwd(), os.path.relpath("actions/Resources/event_details.json"))
        data = pd.read_json(loc)  
        data = pd.DataFrame(data)
        entity_name = "event_name"
        for i in range(len(data["Event_Name"])):
            #############################################
            e_name = "event_name"
            e_value = data["Event_Name"][i]
            buttons.append(
                        {"title": e_value, "payload": "/event_choice "+json.dumps({e_name:e_value})}
                    )
            
        dispatcher.utter_message(response="utter_event_name_details", buttons=buttons, button_type = "vertical")

class ActionTellSpecificEventChoices(Action):
    def name(self) -> Text:
        return "action_tell_specific_event_choices"

    def run(
        self,
        dispatcher: CollectingDispatcher,
        tracker: Tracker,
        domain: DomainDict,
    ) -> None:
        sp_event_name = next(tracker.get_latest_entity_values("event_name"), None)
        if sp_event_name.lower() in ["icebreaker","icebreakers","ice breaker", "ice breakers", "stem"]:
            if sp_event_name.lower() in ["icebreaker","icebreakers","ice breaker", "ice breakers"]:
                buttons = []
                loc = os.path.join(os.getcwd(), os.path.relpath("actions/Resources/ice_breakers_details.json"))
                data = pd.read_json(loc)  
                data = pd.DataFrame(data)
                entity_name = "ice_breaker_club_name"
                for i in range(len(data["Club_Name"])):
                    #############################################
                    e_name = "ice_breaker_club_name_name"
                    e_value = data["Club_Name"][i]
                    buttons.append(
                                {"title": e_value, "payload": "/ice_breaker_club_name "+json.dumps({e_name:e_value})}
                            )
                    
                dispatcher.utter_message(response="utter_event_name_details", buttons=buttons, button_type = "vertical")

            if sp_event_name.lower() in ["stem"]:
                buttons = []
                loc = os.path.join(os.getcwd(), os.path.relpath("actions/Resources/stem_details.json"))
                data = pd.read_json(loc)  
                data = pd.DataFrame(data)
                entity_name = "stem_club_name"
                for i in range(len(data["Club_Name"])):
                    #############################################
                    e_name = "stem_club_name_name"
                    e_value = data["Stem_Event_Name"][i]
                    buttons.append(
                                {"title": e_value, "payload": "/stem_club_name "+json.dumps({e_name:e_value})}
                            )
                    
                dispatcher.utter_message(response="utter_event_name_details", buttons=buttons, button_type = "vertical")
        else:
            dispatcher.utter_message(response="utter_basic_options")

class ActionTellSpecificClubInfo(Action):

    def name(self) -> Text:
        return "action_tell_specific_club_info"

    async def run(
        self, dispatcher, tracker: Tracker, domain: Dict[Text, Any],
    ) -> List[Dict[Text, Any]]:
        sp_event_name = next(tracker.get_latest_entity_values("event_name"), None)
        # club_name = next(tracker.get_latest_entity_values("club_name"), None)
        
        # loc = os.path.join(os.getcwd(), os.path.relpath("actions/Resources/club_details.json"))

        # data = pd.read_json(loc)  
        # data = pd.DataFrame(data)
        # result_data = data.query("Name == @club_name")
        # if not result_data.empty:
        #     msg = "Club Name : "+club_name+"\nClub Lead : "+result_data["Club_Lead"].iloc[0]+ "\nClub President : "+result_data["Club_President"].iloc[0]+"\nClub Description : "+result_data["Club_Description"].iloc[0]+"\nInstagram Handle : https://instagram.com/"+result_data["Insta_handle"].iloc[0]
        # else:
        #     msg = f"The club you are looking for doesn't seem to exist. Could you please check again"
          
        # dispatcher.utter_message(text=msg)
        
        # return []
        if sp_event_name.lower() in ["icebreaker","icebreakers","ice breaker", "ice breakers", "stem"]:
            if sp_event_name.lower() in ["icebreaker","icebreakers","ice breaker", "ice breakers"]:
                club_name = next(tracker.get_latest_entity_values("ice_breaker_club_name"), None)
        
                loc = os.path.join(os.getcwd(), os.path.relpath("actions/Resources/ice_breakers_details.json"))

                data = pd.read_json(loc)  
                data = pd.DataFrame(data)
                result_data = data.query("Club_Name == @club_name")
                if not result_data.empty:
                    msg = "Club Name : "+club_name+"\nFaculty in charge : "+result_data["Faculty_in_charge "].iloc[0]+ "\nClub President : "+result_data["Club_President"].iloc[0]+"\nClub Description : "+result_data["What_is_the_club"].iloc[0]+"\nIce breakers event : " +result_data["Activity_at_event"].iloc[0]
                else:
                    msg = f"The club you are looking for doesn't seem to exist. Could you please check again"
                
                dispatcher.utter_message(text=msg)
                
                return []

            elif sp_event_name.lower() in ["stem"]:
                club_name = next(tracker.get_latest_entity_values("stem_club_name"), None)
        
                loc = os.path.join(os.getcwd(), os.path.relpath("actions/Resources/ice_breakers_details.json"))

                data = pd.read_json(loc)  
                data = pd.DataFrame(data)
                result_data = data.query("Club_Name == @club_name")
                if not result_data.empty:
                    msg = "Club Name : "+club_name+"\nContact : "+result_data["Who_to_contact"].iloc[0]+ "\nEvent details : "+result_data["What_is_the_event"].iloc[0]+"\nWhat's in the event: : "+result_data["Included_in_event"].iloc[0]+"\nTransport : "+result_data["Transport_after_event"].iloc[0]
                else:
                    msg = f"The club you are looking for doesn't seem to exist. Could you please check again"
                
                dispatcher.utter_message(text=msg)
                
                return []
        

class ActionDefaultAskAffirmation(Action):
    """Asks for an affirmation of the intent if NLU threshold is not met."""

    def name(self) -> Text:
        return "action_default_ask_affirmation"

    def __init__(self) -> None:
        import pandas as pd

        self.intent_mappings = pd.read_csv(INTENT_DESCRIPTION_MAPPING_PATH)
        self.intent_mappings.fillna("", inplace=True)
        self.intent_mappings.entities = self.intent_mappings.entities.map(
            lambda entities: {e.strip() for e in entities.split(",")}
        )

    def run(
        self,
        dispatcher: CollectingDispatcher,
        tracker: Tracker,
        domain: DomainDict,
    ) -> List[EventType]:

        intent_ranking = tracker.latest_message.get("intent_ranking", [])
        if len(intent_ranking) > 1:
            diff_intent_confidence = intent_ranking[0].get(
                "confidence"
            ) - intent_ranking[1].get("confidence")
            if diff_intent_confidence < 0.7:
                intent_ranking = intent_ranking[:2]
            else:
                intent_ranking = intent_ranking[:1]

        # for the intent name used to retrieve the button title, we either use
        # the name of the name of the "main" intent, or if it's an intent that triggers
        # the response selector, we use the full retrieval intent name so that we
        # can distinguish between the different sub intents
        first_intent_names = [
            intent.get("name", "")
            if intent.get("name", "") not in ["faq", "chitchat"]
            else tracker.latest_message.get("response_selector")
            .get(intent.get("name", ""))
            .get("ranking")[0]
            .get("intent_response_key")
            for intent in intent_ranking
        ]
        if "nlu_fallback" in first_intent_names:
            first_intent_names.remove("nlu_fallback")
        if "/out_of_scope" in first_intent_names:
            first_intent_names.remove("/out_of_scope")
        if "out_of_scope" in first_intent_names:
            first_intent_names.remove("out_of_scope")

        if len(first_intent_names) > 0:
            message_title = (
                "Sorry, I'm not sure I've understood you correctly 🤔 Do you mean..."
            )

            entities = tracker.latest_message.get("entities", [])
            entities = {e["entity"]: e["value"] for e in entities}

            entities_json = json.dumps(entities)

            buttons = []
            for intent in first_intent_names:
                button_title = self.get_button_title(intent, entities)
                if "/" in intent:
                    # here we use the button title as the payload as well, because you
                    # can't force a response selector sub intent, so we need NLU to parse
                    # that correctly
                    buttons.append({"title": button_title, "payload": button_title})
                else:
                    buttons.append(
                        {"title": button_title, "payload": f"/{intent}{entities_json}"}
                    )

            buttons.append({"title": "Something else", "payload": "/out_of_scope"})

            dispatcher.utter_message(text=message_title, buttons=buttons)
        else:
            message_title = (
                "Sorry, I'm not sure I've understood "
                "you correctly 🤔 Can you please rephrase?"
            )
            dispatcher.utter_message(text=message_title)

        return []

    def get_button_title(self, intent: Text, entities: Dict[Text, Text]) -> Text:
        default_utterance_query = self.intent_mappings.intent == intent
        utterance_query = (self.intent_mappings.entities == entities.keys()) & (
            default_utterance_query
        )

        utterances = self.intent_mappings[utterance_query].button.tolist()

        if len(utterances) > 0:
            button_title = utterances[0]
        else:
            utterances = self.intent_mappings[default_utterance_query].button.tolist()
            button_title = utterances[0] if len(utterances) > 0 else intent

        return button_title.format(**entities)

class ActionDefaultFallback(Action):
    def name(self) -> Text:
        return "action_default_fallback"

    def run(
        self,
        dispatcher: CollectingDispatcher,
        tracker: Tracker,
        domain: DomainDict,
    ) -> List[EventType]:

        # Fallback caused by TwoStageFallbackPolicy
        last_intent = tracker.latest_message["intent"]["name"]
        if last_intent in ["nlu_fallback", USER_INTENT_OUT_OF_SCOPE]:
            return [SlotSet("feedback_value", "negative")]

        # Fallback caused by Core
        else:
            dispatcher.utter_message(response="utter_default")
            return [UserUtteranceReverted()]

class ActionTriggerResponseSelector(Action):
    """Returns the chitchat utterance dependent on the intent"""

    def name(self) -> Text:
        return "action_trigger_response_selector"

    def run(
        self,
        dispatcher: CollectingDispatcher,
        tracker: Tracker,
        domain: Dict[Text, Any],
    ) -> List[EventType]:
        retrieval_intent = tracker.get_slot("retrieval_intent")
        if retrieval_intent:
            dispatcher.utter_message(response=f"utter_{retrieval_intent}")

        return [SlotSet("retrieval_intent", None)]

# class ActionTagFeedback(Action):
#     """Tag a conversation in Rasa X as positive or negative feedback"""

#     def name(self):
#         return "action_tag_feedback"

#     def run(
#         self,
#         dispatcher: CollectingDispatcher,
#         tracker: Tracker,
#         domain: DomainDict,
#     ) -> List[EventType]:

#         feedback = tracker.get_slot("feedback_value")

#         if feedback == "positive":
#             label = '[{"value":"postive feedback","color":"76af3d"}]'
#         elif feedback == "negative":
#             label = '[{"value":"negative feedback","color":"ff0000"}]'
#         else:
#             return []

#         rasax = RasaXAPI()
#         rasax.tag_convo(tracker, label)

#         return []
